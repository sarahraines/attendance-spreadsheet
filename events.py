from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
import re
import copy
from openpyxl.utils import get_column_letter
from titlecase import titlecase
from fuzzywuzzy import process, fuzz
import sys

def new_tup(orgname, year):
  switch = {
    2014: (orgname, 1, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
    2015: (orgname, None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
    2016: (orgname, None, None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
    2017: (orgname, None, None, None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None),
    2018: (orgname, None, None, None, None, 1, None, None, None, None, None, None, None, None, None, None, None, None),
    2019: (orgname, None, None, None, None, None, 1, None, None, None, None, None, None, None, None, None, None, None),
    2020: (orgname, None, None, None, None, None, None, 1, None, None, None, None, None, None, None, None, None, None),
    2021: (orgname, None, None, None, None, None, None, None, 1, None, None, None, None, None, None, None, None, None),
    2022: (orgname, None, None, None, None, None, None, None, None, 1, None, None, None, None, None, None, None, None),
    2023: (orgname, None, None, None, None, None, None, None, None, None, 1, None, None, None, None, None, None, None),
    2024: (orgname, None, None, None, None, None, None, None, None, None, None, 1, None, None, None, None, None, None),
    2025: (orgname, None, None, None, None, None, None, None, None, None, None, None, 1, None, None, None, None, None),
    2026: (orgname, None, None, None, None, None, None, None, None, None, None, None, None, 1, None, None, None, None),
    2027: (orgname, None, None, None, None, None, None, None, None, None, None, None, None, None, 1, None, None, None),
    2028: (orgname, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 1, None, None),
    2029: (orgname, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 1, None),
    2030: (orgname, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 1),                    
  }
  return switch.get(year, ("Invalid year", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None))

def update_row(row, year):
  row[year].value = 1

def row_to_tuple(row):
  return [row[i].value for i in xrange(18)]

if __name__ == '__main__':

  year = input("What year is it? Value must be between 2014-2030. ")

  if year > 2030 or year < 2014:
    sys.exit("The year you entered is not within the defined range (2014-2030).")

  wb = load_workbook('Event Data.xlsx')
  ws = wb['Yearly Participation']

  rows = {}

  f = open('orgs.txt', 'r')

  orgs = {x.strip().decode('utf-8') for x in f.readlines()}
  orgs = [re.sub(r'\([^()]*\)', '', org) for org in orgs]
  orgs = [titlecase(org) for org in orgs]

  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value != None:
      if row[0].value in orgs:
        orgs.remove(row[0].value)
        update_row(row, year-2013)
      else:
        my_tup = process.extractOne(row[0].value, orgs)
        if my_tup[1] >= 87:
          orgs.remove(my_tup[0])
          update_row(row, year-2013)

  for org in orgs: 
    rows[org] = new_tup(org, year)

  for row in rows.values():
    ws.append(row)

  wb.save('Event Data.xlsx')
  
  print 'Spreadsheet updated. Made with <3 by Sarah Raines.'
